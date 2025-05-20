import csv
import json
import random
from datetime import datetime, time, timezone

import openpyxl
from django.contrib import messages
from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.views.decorators.csrf import csrf_exempt
from openpyxl import load_workbook, workbook
from openpyxl.utils import get_column_letter

from apps.attendance.models import Attendance
from apps.fees.models import Fee
from apps.profiles.models import ParentProfile
from apps.settings.models import Setting
from apps.terms.models import AcademicYear, ExaminationSession, Term
import logging

from .forms import VerifyPinForm
from .models import (
    Class,
    Mark,
    StudentProfile,
    StudentTempCreateProfile,
    Subject,
    Department,
    TeacherProfile,
    TeacherTempCreateProfile,
)


# Configure logging
logger = logging.getLogger(__name__)
from .utils import (
    check_student_is_owing,
    check_student_is_repeater,
    create_student_parent,
    create_student_pin,
    format_date,
    generate_random_pin,
    get_student_temp_account,
    get_teacher_temp_account,
    send_account_creation_email,
    send_password_reset_email,
)

User = get_user_model()

from faker import Faker

faker = Faker()
logger = logging.getLogger(__name__)


@login_required
def list_student_view(request):

    queryset = StudentProfile.objects.all()
    classes = Class.objects.all()
    departments = Department.objects.all()
    template_name = "students/list.html"
    context = {
        "section": "student-area",
        "deparments": departments,
        "students": queryset,
        "classes": classes,
    }

    return render(request, template_name, context)


@login_required
def student_detail_view(request, matricule, pkid):
    fee = None
    payment_history = None
    student = get_object_or_404(StudentProfile, matricule=matricule, pkid=pkid)
    extra_payment_history = student.extra_payments.all()
    # get the current academic year
    current_year = AcademicYear.objects.get(is_current=True)
    terms = Term.objects.filter(academic_year=current_year)

    # Get the payment history for the current year
    fees = Fee.objects.filter(student=student, academic_year=current_year)
    payment_history = []
    total_amount_paid = 0
    for fee in fees:
        payment_his = fee.fee_history.all()
        for payment in payment_his:
            payment_history.append(payment)
            total_amount_paid += payment.amount_paid
    # Grab all the mark records associated to this student
    marks = student.student_marks.all()

    # Get all the subjects in the class the student belongs to
    # and those optionally added by the student
    subjects1 = student.current_class.subjects.all()
    # get optoinal subjects for the particular student
    optional_subjects = student.optional_subjects.all()

    distinct_subjects = set(list(subjects1) + list(optional_subjects))

    pin = StudentTempCreateProfile.objects.filter(student=student)
    if pin.exists():
        pin = pin.first()
    else:
        pin = None

    template_name = "students/details.html"
    context = {
        "section": "student-area",
        "student_detial_section": True,
        "student": student,
        "payment_history": payment_history,
        "number_of_payments": len(payment_history),
        "marks": marks,
        "total_payment": total_amount_paid,
        "subjects": distinct_subjects,
        "pin": pin,
        "extra_payments": extra_payment_history,
        "terms": terms,
    }

    return render(request, template_name, context)


@login_required
def add_student_view(request):
    classes = Class.objects.all()
    if request.method == "POST":
        print(request.POST)

        pin = request.POST.get("password_reset_pin")
        parent_fname = request.POST.get("first_name")
        parent_occupation = request.POST.get("parent-occupation")
        parent_phone = request.POST.get("parent-phone")
        parent_email = request.POST.get("parent-email")
        parent_address = request.POST.get("parent-address")
        parent_role = request.POST.get("parent-role")
        st_class = request.POST.get("class")

        # Check if student pin is valid
        # should not be greater or less than 4 in length
        if len(pin) == 4:
            # create temp
            pass
        else:
            messages.error(
                "Password reset pin has incorrect length, should be exactly 4."
            )
            return redirect(reverse("students:student-add"))

        # Get class in which student is part of
        print("for the class", st_class)
        student_class = get_object_or_404(Class, pkid=int(st_class))

        # Create parent for student
        student_parent, created = ParentProfile.objects.get_or_create(
            first_name=parent_fname,
            phone=parent_phone,
            occupation=parent_occupation,
            address=parent_address,
            email=parent_email,
            role=parent_role,
        )

        student_parent.save()

        st_fname = request.POST.get("first_name")
        st_lname = request.POST.get("last_name")
        st_gender = request.POST.get("selected_gender")
        st_email = request.POST.get("email")
        st_dob = request.POST.get("dob")
        st_address = request.POST.get("address")
        st_phone = request.POST.get("phone")
        st_domain = request.POST.get("domain")
        st_image = request.FILES.get("profile_photo")

        # Create the user instance
        # construct a valid date out of the html date
        date_string_from_form = st_dob

        dob = datetime.strptime(date_string_from_form, "%Y-%m-%d").date()
        # Create a specific time
        specific_time = time(8, 51, 2)

        dob_with_time = datetime.combine(dob, specific_time, tzinfo=timezone.utc)

        # Construct faker usernames, this is done assuming that the usernames are not important to the students
        # we avoid using direct fake usernames, so as to avoid unique constrainst failure in the database
        # This is why the random digit is added to the end of every fake username generated by faker
        # faker_username = st_fname
        # attache random digit to the end
        # faker_username + str(random.randint(0, 9))

        if not st_email:
            st_email = faker.email()

        user, created = User.objects.get_or_create(
            username=st_fname,
            first_name=st_fname,
            last_name=st_lname,
            email=st_email,
            is_student=True,
            dob=dob_with_time,
        )
        if created:
            user.save()

        # Create student instance

        student, created = StudentProfile.objects.get_or_create(
            user=user,
            parent=student_parent,
            current_class=student_class,
            gender=st_gender,
            phone_number=st_phone,
            # profile_photo=st_photo,
            address=st_address,
            domain=st_domain,
        )

        if created:

            student.save()
            if st_image:
                student.profile_photo = st_image
            student.save()
            create_student_pin(student, pin)
            send_account_creation_email(request, student.user)
            return redirect(reverse("students:student-list"))
        else:
            messages.error(request, "Student already exist.")
            return redirect(reverse("students:student-list"))
    departments = Department.objects.all()
    template_name = "students/add.html"

    context = {
        "section": "student-area",
        "classes": classes,
        "departments": departments,
    }

    return render(request, template_name, context)


def get_classes_by_department(request):
    department_id = request.GET.get("department_id")
    classes = Class.objects.filter(department_id=department_id).values(
        "pkid", "class_name"
    )
    return JsonResponse(list(classes), safe=False)


@login_required
def edit_student_profile(request, pkid, matricule):
    student = get_object_or_404(StudentProfile, pkid=pkid, matricule=matricule)

    if request.method == "POST":
        st_class = request.POST.get("student_class")
        print("for the class", st_class)
        student_class = get_object_or_404(Class, pkid=int(st_class))

        st_fname = request.POST.get("first_name")
        st_lname = request.POST.get("last_name")
        st_gender = request.POST.get("selected_gender")
        st_email = request.POST.get("email")
        st_dob = request.POST.get("dob")
        st_address = request.POST.get("address")
        st_phone = request.POST.get("phone")
        st_domain = request.POST.get("domain")
        st_image = request.FILES.get("profile_photo")

        # Parse date string from form
        dob = datetime.strptime(st_dob, "%Y-%m-%d").date()
        specific_time = time(8, 51, 2)

        dob_with_time = datetime.combine(dob, specific_time, tzinfo=timezone.utc)
        # <class 'datetime.datetime'>
        # Update user fields
        student.user.first_name = st_fname
        student.user.last_name = st_lname
        student.user.email = st_email
        student.user.dob = dob_with_time

        if st_image:
            student.profile_photo = st_image
        student.user.save()

        # Update student fields
        student.current_class = student_class
        student.gender = st_gender
        student.phone_number = st_phone
        student.address = st_address
        student.domain = st_domain

        student.save()
        return redirect(
            reverse(
                "students:student-detail",
                kwargs={"pkid": student.pkid, "matricule": student.matricule},
            )
        )

    # Get and associate a parent to the request.
    parent = student.parent
    departments = Department.objects.all()
    template_name = "students/edit.html"
    context = {
        "section": "student_area",
        "student_profile": student,
        "dob": str(student.user.dob.date()),
        "classes": Class.objects.all(),
        "departments": departments,
    }

    return render(request, template_name, context)


@login_required
def download_marksheet(request, subject_pkid, class_pkid, *args, **kwargs):
    # Get the class for which the mark sheet needs to be downloaded}|

    klass = get_object_or_404(Class, pkid=class_pkid)
    logger.info(f"IDs: {subject_pkid}-{class_pkid}")
    current_academic_session = AcademicYear.objects.filter(is_current=True).first()
    current_exam_session = ExaminationSession.objects.filter(is_current=True).first()
    if request.method == "POST":

        subject = get_object_or_404(Subject, pkid=subject_pkid)
        # Fetch all the students associated with the class
        students = klass.students.all()
        # Create a new workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "marks"

        # Define column headers
        headers = [
            "Matricule",
            "Full Name",
            "Mark",
        ]

        # Write headers to the first row
        for col_num, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header

        # Write student data to the worksheet
        for row_num, student in enumerate(students, start=2):
            ws.cell(row=row_num, column=1).value = student.matricule
            ws.cell(row=row_num, column=2).value = student.user.get_fullname

        # Set column widths
        for col_num in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_num)
            ws.column_dimensions[column_letter].width = 20

        # Create a response object
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        filename = f"{klass.grade_level}-{klass.class_name}-{subject.name}.xlsx"
        logger.info(
            f"Filename: {filename}, Grade: {klass.grade_level}, ClassName: {klass.class_name}, Subject: {subject.name}"
        )
        response["Content-Disposition"] = f"attachment; filename={filename}"

        # Save the workbook to the response
        wb.save(response)

        return response

    template_name = "students/marks-sheet-download.html"
    context = {"section": "marks-area"}
    return render(request, template_name, context)


def upload_marks111(request, subject_pkid, class_pkid, *args, **kwargs):
    # Get the subject and all the students associated to thesubject from the database

    subject = Subject.objects.get(pkid=subject_pkid)
    klass = Class.objects.get(pkid=class_pkid)
    academic_year = AcademicYear.objects.filter(is_current=True).first()

    terms = Term.objects.filter(academic_year=academic_year)

    sequences = terms.filter(is_current=True).first().examination_sessions.all()

    teacher = request.user.teacher_profile
    # Get data for student marks and updates
    students = klass.students.all()
    # Get all the marks for the given subject, class, and term
    term = Term.objects.get(is_current=True)
    # get the value for the sequences in the term, (Example: 1st, 2nd)
    session_ids = sequences.values_list("pkid", flat=True)
    first_seq_id = session_ids.first()
    second_seq_id = session_ids.last()
    first_session_name = ExaminationSession.objects.get(pkid=first_seq_id)
    second_session_name = ExaminationSession.objects.get(pkid=second_seq_id)

    data = []
    for student in students:
        name = student.user.get_fullname
        fmark = Mark.objects.filter(
            student=student, exam_session__pkid=first_seq_id, subject=subject
        ).first()
        lmark = Mark.objects.filter(
            student=student, exam_session__pkid=second_seq_id, subject=subject
        ).first()
        score1 = 0
        score2 = 0
        if not fmark:
            fmark = 0
        else:
            score1 = fmark.score
        if not lmark:
            lmark = 0
        else:
            score2 = lmark.score

        obj = {
            "student": student,
            "name": name,
            "fmark": fmark,
            "lmark": lmark,
            "score1": score1,
            "score2": score2,
        }
        data.append(obj)

    # for term in terms:
    #     for ex_session in term.examination_sessions.all():
    #         sequences.append(ex_session)

    if request.method == "POST" and request.FILES["marks_file"]:
        setting = Setting.objects.all().first()
        if setting.teacher_can_upload == False:
            messages.error(request, "Can not upload at the moment")
            return redirect(
                reverse(
                    "students:marks-upload",
                    kwargs={"subject_pkid": subject_pkid, "class_pkid": class_pkid},
                )
            )

        marks_file = request.FILES["marks_file"]

        selected_ex_session_id = request.POST.get("selected_ex_session")
        exam_session = ExaminationSession.objects.filter(pkid=selected_ex_session_id)
        if exam_session.exists():
            exam_session = exam_session.first()
        else:
            messages.error(request, "Invalid Exam Session")
            return redirect(
                reverse(
                    "students:marks-upload",
                    kwargs={"subject_pkid": subject_pkid, "class_pkid": class_pkid},
                )
            )

        # decoded_file = marks_file.read().decode('utf-8').splitlines()

        # Get the teacher from the DB
        # teacher = request.user.teacher_profile

        # Get the subject from the database

        wb = load_workbook(filename=marks_file)

        ws = wb["marks"]

        for row in ws.iter_rows(min_row=2, values_only=True):
            student_matricule, subject_name, marks = (
                row[0],
                row[1],
                row[2],
            )
            print(
                f"this is the student information student mat: {student_matricule}, mark: {marks}, subjectname: {subject_name}, subject_id: {subject.name}"
            )
            student = StudentProfile.objects.get(matricule=student_matricule)
            # Check if a mark already exists for this student and subject
            mark, created = Mark.objects.get_or_create(
                student=student, subject=subject, exam_session=exam_session
            )

            mark.teacher = subject.assigned_to

            # Update the score
            if not marks:
                marks = 0
            mark.score = marks
            mark.save()
            print(mark)

        messages.success(
            request,
            f"Marks have been updated for the subject: {subject.name} by: {teacher.user.username} with matricule No: {teacher.matricule}",
        )
        return redirect(reverse("students:marks"))

    template_name = "students/upload-marks.html"
    context = {
        "section": "marks-area",
        "subject": subject,
        "klass": klass,
        "sessions": sequences,
        "year": academic_year,
        "students": students,
        "data": data,
        "first_session_name": first_session_name,
        "second_session_name": second_session_name,
    }

    return render(request, template_name, context)


@login_required
def upload_marks1(request, subject_pkid, class_pkid, *args, **kwargs):
    # Get the subject and all the students associated with the subject from the database
    subject = Subject.objects.get(pkid=subject_pkid)
    klass = Class.objects.get(pkid=class_pkid)
    academic_year = AcademicYear.objects.filter(is_current=True).first()
    terms = Term.objects.filter(academic_year=academic_year)
    sequences = terms.filter(is_current=True).first().examination_sessions.all()
    teacher = request.user.teacher_profile

    # Get all students for the class
    students = klass.students.all()
    term = Term.objects.get(is_current=True)
    session_ids = sequences.values_list("pkid", flat=True)
    first_seq_id = session_ids.first()
    second_seq_id = session_ids.last()
    first_session_name = ExaminationSession.objects.get(pkid=first_seq_id)
    second_session_name = ExaminationSession.objects.get(pkid=second_seq_id)

    # Prepare the data for rendering
    data = []
    for student in students:
        name = student.user.get_fullname
        fmark = Mark.objects.filter(
            student=student, exam_session__pkid=first_seq_id, subject=subject
        ).first()
        lmark = Mark.objects.filter(
            student=student, exam_session__pkid=second_seq_id, subject=subject
        ).first()
        score1 = 0
        score2 = 0
        if fmark:
            score1 = fmark.score
        if lmark:
            score2 = lmark.score

        obj = {
            "student": student,
            "name": name,
            "fmark": fmark,
            "lmark": lmark,
            "score1": score1,
            "score2": score2,
        }
        data.append(obj)

    # Handle the file upload and processing
    if request.method == "POST" and request.FILES.get("marks_file"):
        setting = Setting.objects.first()
        if not setting.teacher_can_upload:
            messages.error(request, "Cannot upload at the moment")
            return redirect(
                reverse(
                    "students:marks-upload",
                    kwargs={"subject_pkid": subject_pkid, "class_pkid": class_pkid},
                )
            )

        marks_file = request.FILES["marks_file"]
        selected_ex_session_id = request.POST.get("selected_ex_session")
        exam_session = ExaminationSession.objects.filter(
            pkid=selected_ex_session_id
        ).first()
        if not exam_session:
            messages.error(request, "Invalid Exam Session")
            return redirect(
                reverse(
                    "students:marks-upload",
                    kwargs={"subject_pkid": subject_pkid, "class_pkid": class_pkid},
                )
            )

        # Retrieve the maximum mark allowed
        max_mark = setting.highest_upload_mark
        min_mark = 0  # Minimum mark is 0

        # Load the workbook and process each row
        wb = load_workbook(filename=marks_file)
        ws = wb["marks"]

        row_errors = []  # Collect errors for each row
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            student_matricule, subject_name, marks = row
            logger.info(
                f"Processing row {idx}: Student mat: {student_matricule}, Mark: {marks}"
            )

            student = StudentProfile.objects.filter(matricule=student_matricule).first()
            if not student:
                row_errors.append(
                    f"Row {idx}: Student with matricule {student_matricule} does not exist."
                )
                continue

            mark, created = Mark.objects.get_or_create(
                student=student, subject=subject, exam_session=exam_session
            )
            mark.teacher = subject.assigned_to

            # Validate the score
            if marks is None:
                marks = 0
            elif marks < min_mark or marks > max_mark:
                row_errors.append(
                    f"Row {idx}: Mark {marks} is out of bounds. Must be between {min_mark} and {max_mark}."
                )
                continue  # Skip updating this row if there's an error

            mark.score = marks
            mark.save()
            logger.info(f"Updated mark for student {student_matricule} to {marks}.")

        # Inform user about errors
        if row_errors:
            error_message = (
                "There were problems with the following rows:\n" + "\n".join(row_errors)
            )
            messages.error(request, error_message)
            logger.error(error_message)
            return redirect(
                reverse(
                    "students:marks-upload",
                    kwargs={"subject_pkid": subject_pkid, "class_pkid": class_pkid},
                )
            )

        messages.success(
            request,
            f"Marks have been updated for the subject: `{subject.name}` by: `{teacher.user.username}` with matricule No: {teacher.matricule}",
        )
        return redirect(reverse("students:marks"))

    # Render the template with context
    template_name = "students/upload-marks.html"
    context = {
        "section": "marks-area",
        "subject": subject,
        "klass": klass,
        "sessions": sequences,
        "year": academic_year,
        "students": students,
        "data": data,
        "first_session_name": first_session_name,
        "second_session_name": second_session_name,
    }

    return render(request, template_name, context)


@login_required
def teacher_modify_student_mark(
    request, subject_pkid, student_pkid, f_session, l_session
):
    """
    Allows to teachers to modify students marks for a given course
    """
    # get the values from the form

    if request.method == "POST":
        f_session_mark = request.POST.get("fmark")
        l_session_mark = request.POST.get("lmark")
        session1 = ExaminationSession.objects.get(pkid=f_session)
        session2 = ExaminationSession.objects.get(pkid=l_session)
        subject = Subject.objects.get(pkid=subject_pkid)

        student = StudentProfile.objects.get(pkid=student_pkid)
        # Create mark instances
        student_mark1, created = Mark.objects.update_or_create(
            student=student, exam_session=session1, subject=subject
        )
        student_mark1.score = f_session_mark
        student_mark1.save()

        student_mark2, created = Mark.objects.update_or_create(
            student=student, exam_session=session2, subject=subject
        )
        student_mark2.score = l_session_mark
        student_mark2.save()
        messages.success(request, "Marks updated successfully")
        return redirect(
            reverse(
                "students:marks-upload",
                kwargs={
                    "class_pkid": student.current_class.pkid,
                    "subject_pkid": subject.pkid,
                },
            )
        )

    messages.error(request, "request")
    return redirect(reverse("students:marks"))


@login_required
def fill_student_marks(request, subject_pkid, class_pkid):
    if request.method == "POST":
        selecte_session_id = request.POST.get("selected_ex_session")
        sessions = ExaminationSession.objects.filter(pkid=selecte_session_id)
        klasses = Class.objects.filter(pkid=class_pkid)
        subjects = Subject.objects.filter(pkid=subject_pkid)

        if subjects.exists():
            subject = subjects.first()
        else:
            messages(request, "Invalid Class")
            return redirect(reverse("students:marks"))

        if klasses.exists():
            klass = klasses.first()
        else:
            messages(request, "Invalid Class")
            return redirect(reverse("students:marks"))
        if sessions.exists():
            session = sessions.first()
        else:
            messages(request, "Invalid session")
            return redirect(reverse("students:marks"))

        # Get all the students in the class
        students = klass.students.all()
        template_name = "students/fill-marks.html"
        context = {
            "section": "marks",
            "session": session,
            "students": students,
            "subject": subject,
        }
        return render(request, template_name, context)


@csrf_exempt
def update_fill_marks(request):
    if request.method == "POST":
        data = json.loads(request.body)
        student_id = data.get("student_id")
        subject = data.get("subject")
        mark = data.get("mark")
        exam_session_pkid = data.get("session")

        print(data)

        teacher = request.user.teacher_profile

        try:
            student = StudentProfile.objects.get(pkid=student_id)
            subject = Subject.objects.get(pkid=subject)
            session = ExaminationSession.objects.get(pkid=exam_session_pkid)
            subject_mark, created = Mark.objects.update_or_create(
                student=student,
                subject=subject,
                exam_session=session,
                teacher=teacher,
                defaults={"score": mark},
            )
            return JsonResponse({"success": True})
        except StudentProfile.DoesNotExist:
            return JsonResponse({"success": False, "error": "Student not found"})

    return JsonResponse({"success": False, "error": "Invalid request"})


@login_required
def marks(request):
    user = request.user
    classes = []
    if user.is_teacher or user.is_admin:
        if user.is_teacher:
            # get all assigned subject to the current teacher.
            teacher = user.teacher_profile
            assigned_subjects = Subject.objects.filter(assigned_to=teacher)

            for sub in assigned_subjects:
                if sub.klass == None:
                    continue
                classes.append({"klass": sub.klass, "subject": sub})

            print(classes)

        if len(classes) < 1:
            classes = None

        template_name = "students/marks.html"
        context = {
            "section": "marks-area",
            "classes": classes,
            "subjects": Subject.objects.all(),
        }

        return render(request, template_name, context)
    else:
        messages.error(request, "Unauthorized..")
        return redirect(reverse("users:user-login"))


def list_student_record(request, pkid, matricule, *args, **kwargs):

    # Grab all the mark records associated to this student
    student = get_object_or_404(StudentProfile, pkid=pkid, matricule=matricule)
    current_year = AcademicYear.objects.filter(is_current=True).first()
    marks = student.student_marks.filter(exam_session__term__academic_year=current_year)

    template_name = "students/academic-records.html"

    context = {"section": "student-area", "marks": marks, "student": student}

    return render(request, template_name, context)


def add_optional_subjects_to_student(
    request, student_pkid, student_matricule, *args, **kwargs
):

    student = get_object_or_404(
        StudentProfile, pkid=student_pkid, matricule=student_matricule
    )
    # Get all the subjects in the class the student belongs to
    # and those optionally added by the student
    subjects1 = student.current_class.subjects.all()
    # get optoinal subjects for the particular student
    optional_subjects = student.optional_subjects.all()

    distinct_subjects = set(list(subjects1) + list(optional_subjects))

    subjects = Subject.objects.all()

    if request.method == "POST":
        print("This are the selected subjects", request.body)
        data = json.loads(request.body)

        selected_subjects_ids = []
        for subject in data["selectedSubjects"]:
            print(subject["pkid"])
            selected_subjects_ids.append(subject.get("pkid"))
        print(selected_subjects_ids)
        # Flush out all the subjects that exist in the class and reset
        print(
            "we are inside the loop", len(optional_subjects), len(selected_subjects_ids)
        )
        if len(optional_subjects) > 0 and len(selected_subjects_ids) > 0:
            for subject in optional_subjects:
                print("This is the subject", subject)
                student.optional_subjects.remove(subject)
                student.save()

        print(student.optional_subjects.all())

        # Reassign the subjects to the klass instance
        # Ge throught the incoming ids, and get the subjects associated the with the pkids
        for pkid in selected_subjects_ids:
            sub = Subject.objects.filter(pkid=pkid).first()
            if sub:
                student.optional_subjects.add(sub)
                student.save()

        return JsonResponse({"message": "updated."})

    template_name = "students/add-optional-subjects.html"
    context = {
        "selected_subjects": distinct_subjects,
        "unselected_subjects": subjects,
        "student": student,
    }

    return render(request, template_name, context)


@login_required
def download_class_list(request, *args, **kwargs):
    # Get the class
    if request.method == "POST":
        class_pkid = request.POST.get("selected_class_id")
        if class_pkid == "all":
            print("Yes, pkid is all")
            students = StudentProfile.objects.all()
        else:
            # Check if there is a class given class_id
            classes = Class.objects.filter(pkid=class_pkid)
            if classes.exists():
                klass = classes.first()
            else:
                return redirect(reverse("students:student-list"))

            students = StudentProfile.objects.filter(current_class=klass)

        wb = openpyxl.Workbook()
        ws = wb.active
        if class_pkid == "all":
            ws.title = f"student-list"
            filename = "student-list.xlsx"
        else:
            ws.title = f"{klass.grade_level}-{klass.class_name}"
            filename = f"{klass.grade_level}-{klass.class_name}-class-list.xlsx"

        headers = [
            "Matricule",
            "Fullname",
            "DOB",
            "Gender",
            "IS Owing",
            "Is Repeater",
            "Class",
        ]

        # Write headers to the first row
        for col_num, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header

        # Write student data to the worksheet
        for row_num, student in enumerate(students, start=2):
            ws.cell(row=row_num, column=1).value = student.matricule
            ws.cell(row=row_num, column=2).value = student.user.get_fullname
            ws.cell(row=row_num, column=3).value = student.user.dob.year
            # ws.cell(row=row_num, column=4).value = student.parent.first_name
            # ws.cell(row=row_num, column=4).value = current_academic_session.name
            ws.cell(row=row_num, column=4).value = student.gender
            ws.cell(row=row_num, column=5).value = check_student_is_owing(student.pkid)
            ws.cell(row=row_num, column=6).value = check_student_is_repeater(
                student_pkid=student.pkid
            )
            ws.cell(row=row_num, column=7).value = student.current_class.get_class_name

        # Set column widths
        for col_num in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_num)
            ws.column_dimensions[column_letter].width = 20

        # Create a response object
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f"attachment; filename={filename}"

        # Save the workbook to the response
        wb.save(response)

        return response
    return redirect(reverse("students:student-list"))


@login_required
def upload_students_from_file(request, *args, **kwargs):
    """
    View to upload students from an Excel file and save them to the database.
    """
    classes = Class.objects.all()

    if request.method == "POST" and request.FILES.get("students_file"):
        students_file = request.FILES["students_file"]

        try:
            wb = load_workbook(filename=students_file)
            ws = wb.active
        except Exception as e:
            messages.error(request, f"Error reading the file: {e}")
            return redirect(reverse("students:upload-students-from-file"))

        for row_number, row in enumerate(
            ws.iter_rows(min_row=2, values_only=True), start=2
        ):
            (
                first_name,
                last_name,
                email,
                dob,
                gender,
                phone,
                address,
                pob,
                class_code,
                specialty,
                parent_name,
                parent_phone,
                parent_address,
            ) = row[:13]

            if not all([first_name, last_name, gender, phone]):
                messages.error(
                    request, f"Invalid file, missing information at row {row_number}"
                )
                continue  # Skip the row but continue processing the rest

            dob = format_date(dob)

            if not email:
                email = f"{first_name}{random.randint(1, 1000)}@gmail.com"

            try:
                user, created = User.objects.get_or_create(
                    username=f"{first_name}_{last_name}",
                    first_name=first_name,
                    last_name=last_name,
                    email=email,
                )
                if created:
                    user.dob = dob
                    user.save()
                    logger.info(f"User {user.username} created")

                gender = "Male" if gender.lower() in ["m", "male"] else "Female"

                try:
                    klass = Class.objects.get(class_code=class_code)
                except Class.DoesNotExist:
                    messages.error(request, f"Class not found at row {row_number}")
                    klass = None
                    # Skip the row but continue processing the rest

                student, created = StudentProfile.objects.get_or_create(
                    user=user,
                    defaults={
                        "gender": gender,
                        "phone_number": str(phone),
                        "address": address or faker.address(),
                        "pob": pob,
                        "specialty": specialty,
                    },
                )

                if created:
                    if klass:
                        student.current_class = klass
                    student.save()
                    user.is_student = True
                    user.save()
                    create_student_parent(
                        parent_name, parent_phone, parent_address, student
                    )
                    send_account_creation_email(request, user)
                else:
                    student.current_class = klass
                    student.save()
                    logger.warning(
                        f"Student profile for user {user.username} already exists"
                    )
            except Exception as e:
                logger.error(f"Error processing row {row_number}: {e}")
                messages.error(
                    request, f"There was an error processing row {row_number}."
                )
                continue  # Skip the row but continue processing the rest

        messages.success(request, "Students have been uploaded successfully.")
        return redirect(reverse("students:upload-students-from-file"))

    context = {
        "section": "marks-area",
        "classes": classes,
    }

    return render(request, "students/upload-students.html", context)


@login_required
def download_sample_student_file(request, *args, **kwargs):
    """
    View to download a sample Excel file for student uploads.
    """

    # Create a new workbook and select the active worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Student Upload Sample File"

    # Define the headers for the columns
    headers = [
        "First Name",
        "Last Name",
        "Email",
        "Date Of Birth",
        "Gender",
        "Phone Number",
        "Address",
        "Place of Birth",  # Added field
        "Class",  # Added field
        "Specialty",  # Added field
        "Parent Name",  # Added field
        "Phone Number",  # Added field
        "Parent Address",  # Added field
    ]

    # Write headers to the first row
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header

    # Set column widths to enhance readability
    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)
        ws.column_dimensions[column_letter].width = 20

    # Create an HTTP response object with the correct content type for Excel files
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    # Set the content disposition to attachment to trigger file download
    response["Content-Disposition"] = (
        "attachment; filename=student_upload_sample_file.xlsx"
    )

    # Save the workbook to the response object
    wb.save(response)

    return response


def verify_student_pin(request):
    if request.method == "POST":
        pin = request.POST.get("pin")
        email = request.POST.get("email")
        # Check if student exists with this email
        students = StudentProfile.objects.filter(user__email=email)
        if students.exists():
            student = students.first()
        else:
            messages.error(request, "No student found with given email.")
            return redirect(reverse("users:user-login"))

        # Check if student temp profile exist with this information
        student_temp = get_student_temp_account(pin, email, student)

        if student_temp:
            # go ahead and send them the email to perform a reset
            send_password_reset_email(request, student.user, student_temp)
            return redirect(reverse("users:password_reset_done"))
        else:
            messages.error(request, "No pending setup for the given account")
            return redirect(reverse("users:user-login"))
    else:
        template_name = "accounts/verifypin.html"
        form = VerifyPinForm()
        context = {"form": form}

        return render(request, template_name, context)


from apps.announcements.models import Announcement, Category, Event


def student_dashboard(request):
    print(request.user.username)
    announcements = Announcement.objects.filter(visible_to_students=True)

    # grap the current user from the request
    user = request.user
    # check if the current user is of a student type
    if isinstance(user.student_profile, StudentProfile):
        student = user.student_profile
    else:
        messages.error(request, "Not a valid student instance")
        template_name = "base.html"
        context = {}
        return render(request, template_name, context)

    courses = student.get_all_subjects()
    total_courses = len(courses)

    # calculate fee percentage of fee payment
    fee_percentage_paid = 100  # to be revisited.

    passed_cources = []

    term = Term.objects.get(is_current=True)

    current_year = AcademicYear.objects.get(is_current=True)
    # Get the payment history for the current year
    fees = Fee.objects.filter(student=student, academic_year=current_year)
    payment_history = None

    if fees.exists:
        fee = fees.first()
        payment_history = student.payment_history.filter(fee=fee)
        percentage = None

    # get all the exam sessions ids so far for that term
    examination_session_ids = ExaminationSession.objects.filter(term=term).values_list(
        "pkid", flat=True
    )
    # Grab all the mark records associated to this student for that year.
    marks = student.student_marks.filter(exam_session__pkid__in=examination_session_ids)

    pass_courses_count = marks.filter(score__gte=10).count()

    # get all evernts that have to with the students
    events = Event.objects.filter(visible_to_students=True)

    class_enrolment = StudentProfile.objects.filter(current_class=student.current_class)

    # get present absent data
    student_attendance_present = Attendance.objects.filter(
        student=student, is_present=True
    )
    student_attendance_absent = Attendance.objects.filter(
        student=student, is_present=False
    )

    data_absent_list = []
    data_present_list = []
    data_name_list = []
    for sub in courses:
        data_name_list.append(sub.name)
        # get the sum of all attendances for this subject
        total_attendance_per_sub_present = student_attendance_present.filter(
            subject=sub, is_present=True
        ).count()
        total_attendance_per_sub_absent = student_attendance_absent.filter(
            subject=sub, is_present=False
        ).count()
        data_present_list.append(total_attendance_per_sub_present)
        data_absent_list.append(total_attendance_per_sub_absent)

    total_present = student_attendance_present.count()
    total_absent = student_attendance_absent.count()
    total = total_present + total_absent
    total = total if total else 1
    percent_present = round(((total_present / total) * 100), 2)
    percent_absent = round(((total_absent / total) * 100), 2)
    template_name = "dashboards/student/student-dashboard.html"

    context = {
        "section": "student-dashboard",
        "announcements": announcements,
        "student": student,
        "total_pass_courses": pass_courses_count,
        "total_cources_writen": marks.count(),
        "total_courses": total_courses,
        "payment_history": payment_history,
        "events": events,
        "class": student.current_class,
        "class_enrolment": class_enrolment.count(),
        "fee_percentage_paid": fee_percentage_paid,
        "max_fee_percent": 0,
        # chat information
        "percent_present": percent_present,
        "percent_absent": percent_absent,
        "data_name": data_name_list,
        "data_absent": data_absent_list,
        "data_present": data_present_list,
    }

    print(context)
    return render(request, template_name, context)


def edit_student_marks(request, mark_pkid, student_pkid):
    if request.method == "POST":
        score = request.POST.get("mark")
        mark = Mark.objects.filter(pkid=mark_pkid)
        students = StudentProfile.objects.filter(pkid=student_pkid)
        if students.exists():
            student = students.first()
            if mark.exists():
                mark = mark.first()
                mark.score = score
                mark.save()

                messages.success(request, "Mark updated successfully")
                return redirect(
                    reverse(
                        "students:student-academic-record",
                        kwargs={"pkid": student.pkid, "matricule": student.matricule},
                    )
                )

            messages.error(request, "Invalid mark ID")
            return redirect(
                reverse(
                    "students:student-academic-record",
                    kwargs={"pkid": student.pkid, "matricule": student.matricule},
                )
            )
        messages.error(request, "something went wrong.")
        return redirect(reverse("students:student-list"))
    messages.error(request, "Invlid Request.")
    return redirect(reverse("students:student-list"))
