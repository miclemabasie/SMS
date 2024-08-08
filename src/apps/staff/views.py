import json
import logging
from django.contrib import messages
from django.contrib.auth import get_user_model, logout
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from openpyxl import load_workbook
from django.conf import settings

from django.template.loader import render_to_string
import openpyxl

from apps.scelery.tasks import send_feedback_mail


from apps.attendance.models import Attendance
from apps.profiles.models import ParentProfile
from apps.settings.models import Setting
from apps.students.models import Class, Mark, StudentProfile, Subject, TeacherProfile
from apps.terms.models import AcademicYear, ExaminationSession, Term
from openpyxl.utils import get_column_letter
from .models import AdminProfile

User = get_user_model()

logger = logging.getLogger(__name__)


@login_required
def admin_dashboard(request):

    students = StudentProfile.objects.all().count()
    teachers = TeacherProfile.objects.all().count()
    parents = ParentProfile.objects.all().count()
    classes = Class.objects.all().count()

    # construct data for the header
    header = {
        "students": students,
        "teachers": teachers,
        "parents": parents,
        "classes": classes,
    }

    # Generate data about school attendance
    attendance_list_tuples = []
    student_paid_list_tuples = []
    class_list = []
    student_total_list_perclass = []
    for klass in Class.objects.all():
        class_list.append(klass.get_class_name)
        # got through all the students for a given klass
        students = StudentProfile.objects.filter(current_class=klass)
        total_present = 0
        total_absent = 0
        total_students = 0
        total_student_paid_per_class = 0
        for std in students:
            total_students += 1
            # calculate percentage of students who have paid fieds in the class
            if not std.is_owing:
                total_student_paid_per_class += 1
            print(std)
            # get all the attendance record of every student
            std_attendance_present = Attendance.objects.filter(
                student=std, is_present=True
            )
            std_attendance_absent = Attendance.objects.filter(
                student=std, is_present=False
            )
            print(std_attendance_present)
            print(std_attendance_absent)
            total_present += std_attendance_present.count()
            total_absent += std_attendance_absent.count()
        attendance_list_tuples.append((total_present, total_absent))
        student_total_list_perclass.append(total_students)
        student_paid_list_tuples.append((total_students, total_student_paid_per_class))

    """
    this is example output of the above operation
    ['Form 4-Science', 'Form 5-Kyle Miles', 'Form 6-S2', 'Form 7-S1', 'form1-Form 1 a', 'Form 4-Arts'] [(0, 0), (0, 0), (7, 1), (0, 0), (0, 0), (0, 0)]
    """
    print("new testing", student_paid_list_tuples)
    # Calculate percentage of attendance
    attendance_list = []
    for att in attendance_list_tuples:
        # sub absent and present to get the total
        total_present = att[0]
        total_absent = att[1]
        total = total_present + total_absent
        if total > 0:
            # calculate percentage of attendance
            att_percentage = (total_present / total) * 100
            attendance_list.append(att_percentage)
        else:
            attendance_list.append(0)

    # Calculate percentage of students who have paid fees
    paid_list = []
    for rec in student_paid_list_tuples:
        # sub absent and present to get the total
        total_students = rec[0]
        total_paid_students = rec[1]
        if total_students > 0:
            # calculate percentage of attendance
            att_percentage = round(((total_paid_students / total_students) * 100), 2)
            paid_list.append(att_percentage)
        else:
            paid_list.append(0)

    print(paid_list)

    template_name = "dashboards/staff/dashboard.html"
    context = {
        "section": "admin-area",
        "header": header,
        "subject_list": class_list,
        "class_name_list": class_list,
        "paid_list": paid_list,
        "attendance_list": attendance_list,
        "student_count_per_class": student_total_list_perclass,
    }

    return render(request, template_name, context)


# Subjects
@login_required
def list_all_subjects(request):
    subjects = Subject.objects.all()
    print(subjects)
    template_name = "subjects/subject-list.html"
    context = {
        "section": "subjects-area",
        "subjects": subjects,
    }

    return render(request, template_name, context)


@login_required
def delete_subject(request, pkid, *args, **kwargs):
    subject = get_object_or_404(Subject, pkid=pkid)

    subject.delete()

    return redirect(reverse("staff:subjects"))


@login_required
def add_subject_view(request, *args, **kwargs):
    if request.method == "POST":
        # Extract form data
        subject_name = request.POST.get("subject_name")
        subject_code = request.POST.get("subject_code")
        subject_coeff = request.POST.get("subject_coeff")

        # Create a subject instance
        subject = Subject.objects.create(
            name=subject_name, code=subject_code, coef=subject_coeff
        )
        subject.save()
        messages.success(request, "Subject added.")

        return redirect(reverse("staff:subjects"))
    return redirect(reverse("staff:subjects"))


@login_required
def download_sample_subject_file(request, *args, **kwargs):
    """
    View to download a sample Excel file for subject uploads.
    """
    # Create a new workbook and select the active worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Subject Upload Sample File"

    # Define the headers for the columns
    headers = [
        "Subject Name",
        "Subject Code",
        "Subject Coefficient",
    ]

    # Write headers to the first row
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header

    # Set column widths to enhance readability
    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)
        ws.column_dimensions[column_letter].width = 20

    # Create an HTTP response object with the correct content type for Excel files
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    # Set the content disposition to attachment to trigger file download
    response["Content-Disposition"] = (
        "attachment; filename=subject_upload_sample_file.xlsx"
    )

    # Save the workbook to the response object
    wb.save(response)

    return response


@login_required
def upload_subjects_from_file(request, *args, **kwargs):
    """
    View to handle the file upload for subject data.
    """
    if request.method == "POST" and request.FILES.get("subjects_file"):
        # Get the uploaded file from the request
        subjects_file = request.FILES["subjects_file"]

        try:
            # Load the uploaded workbook
            wb = load_workbook(filename=subjects_file)
            # Get the first sheet from the workbook
            ws = wb.active
        except Exception as e:
            # If there's an error reading the file, show an error message and redirect
            messages.error(request, f"Error reading the file: {e}")
            return redirect(reverse("staff:subjects"))

        # Iterate through each row in the worksheet starting from the second row
        for row_num, row in enumerate(
            ws.iter_rows(min_row=2, values_only=True), start=2
        ):
            # Extract values from the row
            subject_name, subject_code, subject_coeff = row[:3]

            # Check if essential fields are present
            if not subject_name or not subject_code:
                # If any essential field is missing, show an error message and redirect
                messages.error(
                    request,
                    f"Invalid file, missing essential information on row {row_num}",
                )
                return redirect(reverse("staff:subjects"))

            try:
                # Create or update the Subject object
                subject, created = Subject.objects.update_or_create(
                    code=subject_code,  # Use subject_code to look up existing records
                    defaults={
                        "name": subject_name,
                        "coef": subject_coeff,
                    },
                )

                if created:
                    logger.info(f"Subject {subject.name} created")
                else:
                    logger.info(f"Subject {subject.name} updated")

            except Exception as e:
                # Log any errors that occur and show an error message
                logger.error(f"Error processing row {row_num}: {e}")
                messages.error(request, f"Error processing row {row_num}.")
                return redirect(reverse("staff:subjects"))

        # Show a success message after processing all rows and redirect
        messages.success(request, "Subjects have been uploaded successfully.")
        return redirect(reverse("staff:subjects"))

    # Render the form template for uploading subjects
    context = {"section": "subject-area"}
    return redirect(reverse("staff:subjects"))


@login_required
def edit_subject_view(request, pkid, *args, **kwargs):
    subject = get_object_or_404(Subject, pkid=pkid)
    if request.method == "POST":
        # Extract form data
        subject_name = request.POST.get("subject_name")
        subject_code = request.POST.get("subject_code")
        subject_coeff = request.POST.get("subject_coeff")

        # Update subject instance
        subject.name = subject_name
        subject.code = subject_code
        subject.coef = subject_coeff

        subject.save()
        messages.success(request, "Subject Updated succesfully.")

        return redirect(reverse("staff:subjects"))
    return redirect(reverse("staff:subjects"))


@login_required
def assign_subject_to_classes(request, pkid):  # takes in the class pKID
    subjects = Subject.objects.all()
    klass = get_object_or_404(Class, pkid=pkid)

    selected_subjects = Subject.objects.filter(klass=klass)

    unselected_subjects = Subject.objects.filter(has_class=False)

    if request.method == "POST":
        print("This are the selected subjects", request.body)
        data = json.loads(request.body)

        selected_subjects_ids = []
        for subject in data["selectedSubjects"]:
            print(subject["pkid"])
            selected_subjects_ids.append(subject.get("pkid"))

        # Flush out all the subjects that exist in the class and reset
        print("selected subjects", selected_subjects)
        print("unselected subjects", unselected_subjects)

        if len(selected_subjects) > 0:
            for subject in selected_subjects:
                print("This is the subject", subject)
                subject.klass = None
                subject.has_class = False
                subject.save()

        # Reassign the subjects to the klass instance
        # Ge throught the incoming ids, and get the subjects associated the with the pkids
        for pkid in selected_subjects_ids:
            sub = Subject.objects.filter(pkid=pkid).first()
            if sub:
                sub.klass = klass
                sub.has_class = True
                sub.save()

        return JsonResponse({"message": "updated."})

    template_name = "subjects/subject-assign.html"

    context = {
        "section": "subjects",
        "class": klass,
        "unselected_subjects": unselected_subjects,
        "selected_subjects": selected_subjects,
    }

    return render(request, template_name, context)


@login_required
def create_staff(request):
    if request.method == "POST":
        first_name = request.POST.get("first_name")
        last_name = request.POST.get("last_name")
        # subject = request.POST.get("subject")
        gender = request.POST.get("gender")
        phone = request.POST.get("phone")
        dob = request.POST.get("dob")
        address = request.POST.get("address")
        location = request.POST.get("location")
        country = request.POST.get("country")
        photo = request.FILES.get("photo")
        # pin = request.POST.get("pin")
        remark = request.POST.get("remark")
        username = request.POST.get("username")
        email = request.POST.get("email")
        password = request.POST.get("password")
        role = request.POST.get("selected_role")

        # Permissions
        generate_reports = request.POST.get("generate_reports") == "on"
        manage_students = request.POST.get("manage_students") == "on"
        manage_teachers = request.POST.get("manage_teachers") == "on"
        manage_admins = request.POST.get("manage_admins") == "on"
        manage_sessions = request.POST.get("manage_sessions") == "on"
        manage_subjects = request.POST.get("manage_subjects") == "on"

        print(
            "This are different permissions",
            generate_reports,
            manage_students,
            manage_teachers,
            manage_admins,
            manage_sessions,
            manage_subjects,
        )

        # Create User object
        user = User.objects.create_user(
            email=email,
            username=username,
            first_name=first_name,
            last_name=last_name,
            password=password,
            dob=dob,
        )

        user.is_admin = True
        user.is_teacher = True  # the functionality of this line has not been tested.
        user.save()

        # Create AdminProfile object

        admin_profile = AdminProfile(
            user=user,
            gender=gender,
            phone_number=phone,
            country=country,
            location=location,
            address=address,
            role=role,
            remark=remark,
            generate_reports=generate_reports,
            manage_students=manage_students,
            manage_teachers=manage_teachers,
            manage_admins=manage_admins,
            manage_sessions=manage_sessions,
            manage_subjects=manage_subjects,
        )
        admin_profile.save()
        if photo:
            admin_profile.photo = photo
            admin_profile.save()
        # send email to admin
        subject = "Account Creationg Successful."
        from_email = settings.DEFAULT_FROM_EMAIL
        recipient_list = [email]
        email_context = {"name": user.get_fullname}
        html_content = render_to_string("emails/testing.html", email_context)
        # call the Celery task
        send_feedback_mail.delay(subject, "", from_email, recipient_list, html_content)

    template_name = "staff/create-admin.html"
    context = {
        "section": "admin",
    }

    return render(request, template_name, context)


@login_required
def list_admin(request):
    # get all the admins from the system

    admins = AdminProfile.objects.all()

    template_name = "staff/list-admin.html"

    context = {
        "admins": admins,
    }

    return render(request, template_name, context)


@login_required
def mark_list_admin_view(request):
    """
    This view lists out all the classes so the admin can select a specific class to upload marks.
    """
    # Check if usertype is administrator
    if not request.user.is_admin:
        # logout and redirect to login page
        logout(request)
        return redirect("users:user-login")
    # get all the classes in the database
    classes = Class.objects.all()

    template_name = "staff/marks-class-list.html"
    context = {
        "section": "marks",
        "classes": classes,
    }

    return render(request, template_name, context)


@login_required
def subject_marks_list(request, class_pkid):
    """
    Lists out all the subjects in a given class so an administrator can upload marks for a given subject.
    """
    academic_year = AcademicYear.objects.filter(is_current=True).first()

    terms = Term.objects.filter(academic_year=academic_year)
    sequences = terms.filter(is_current=True).first().examination_sessions.all()
    # get the class
    classes = Class.objects.filter(pkid=class_pkid)
    if classes.exists():
        klass = classes.first()
    else:
        messages.error(request, "Class Not found.")
        return redirect(reverse("staff:admin-marks"))

    # get all the courses for this class.
    subjects = klass.subjects.all()

    template_name = "staff/subject-marks-list.html"
    context = {
        "section": "marks",
        "klass": klass,
        "subjects": subjects,
        "sessions": sequences,
    }

    return render(request, template_name, context)


@login_required
def staff_upload_marks(request, subject_pkid, class_pkid, *args, **kwargs):
    # Get the subject and all the students associated to thesubject from the database

    subject = Subject.objects.get(pkid=subject_pkid)
    klass = Class.objects.get(pkid=class_pkid)
    academic_year = AcademicYear.objects.filter(is_current=True).first()

    # terms = Term.objects.filter(academic_year=academic_year)

    # sequences = terms.filter(is_current=True).first().examination_sessions.all()

    # teacher = request.user.teacher_profile
    # # Get data for student marks and updates
    # students = klass.students.all()
    # # Get all the marks for the given subject, class, and term
    # term = Term.objects.get(is_current=True)
    # # get the value for the sequences in the term, (Example: 1st, 2nd)
    # session_ids = sequences.values_list("pkid", flat=True)
    # first_seq_id = session_ids.first()
    # second_seq_id = session_ids.last()
    # first_session_name = ExaminationSession.objects.get(pkid=first_seq_id)
    # second_session_name = ExaminationSession.objects.get(pkid=second_seq_id)

    # data = []
    # for student in students:
    #     name = student.user.get_fullname
    #     fmark = Mark.objects.filter(
    #         student=student, exam_session__pkid=first_seq_id, subject=subject
    #     ).first()
    #     lmark = Mark.objects.filter(
    #         student=student, exam_session__pkid=second_seq_id, subject=subject
    #     ).first()
    #     score1 = 0
    #     score2 = 0
    #     if not fmark:
    #         fmark = 0
    #     else:
    #         score1 = fmark.score
    #     if not lmark:
    #         lmark = 0
    #     else:
    #         score2 = lmark.score

    #     obj = {
    #         "student": student,
    #         "name": name,
    #         "fmark": fmark,
    #         "lmark": lmark,
    #         "score1": score1,
    #         "score2": score2,
    #     }
    #     data.append(obj)

    # # for term in terms:
    # #     for ex_session in term.examination_sessions.all():
    # #         sequences.append(ex_session)

    if request.method == "POST" and request.FILES["marks_file"]:
        # setting = Setting.objects.all().first()
        # if not setting.teacher_can_upload:
        #     messages.error(request, "Can not upload at the moment")
        #     return redirect(
        #         reverse(
        #             "students:marks-upload",
        #             kwargs={"subject_pkid": subject_pkid, "class_pkid": class_pkid},
        #         )
        #     )

        marks_file = request.FILES["marks_file"]
        subject_name = str(subject.name).lower()
        class_name = str(klass.class_name).lower()
        file_name = marks_file.name.lower()

        if subject_name not in file_name:
            # Handle the error
            error_message = f"The uploaded file name does not match the subject name '{subject.name}'."
            messages.error(request, error_message)
            return redirect(
                reverse(
                    "staff:subject-marks-list",
                    kwargs={"class_pkid": klass.pkid},
                )
            )

        if class_name not in file_name:
            # Handle the error
            error_message = f"The uploaded file name does not match the class name '{klass.class_name}'."
            messages.error(request, error_message)
            return redirect(
                reverse(
                    "staff:subject-marks-list",
                    kwargs={"class_pkid": klass.pkid},
                )
            )

        selected_ex_session_id = request.POST.get("selected_ex_session")
        exam_session = ExaminationSession.objects.filter(pkid=selected_ex_session_id)
        if exam_session.exists():
            exam_session = exam_session.first()
        else:
            messages.error(request, "Invalid Exam Session")
            return redirect(
                reverse(
                    "staff:subject-marks-list",
                    kwargs={"class_pkid": klass.pkid},
                )
            )

        # decoded_file = marks_file.read().decode('utf-8').splitlines()

        # Get the teacher from the DB
        # teacher = request.user.teacher_profile

        # Get the subject from the database

        wb = load_workbook(filename=marks_file)

        ws = wb["marks"]

        for row in ws.iter_rows(min_row=2, values_only=True):
            student_matricule, subject_name, marks = (
                row[0],
                row[1],
                row[2],
            )
            print(
                f"this is the student information student mat: {student_matricule}, mark: {marks}, subjectname: {subject_name}, subject_id: {subject.name}"
            )
            student = StudentProfile.objects.get(matricule=student_matricule)
            # Check if a mark already exists for this student and subject
            mark, created = Mark.objects.get_or_create(
                student=student, subject=subject, exam_session=exam_session
            )

            mark.teacher = subject.assigned_to

            # Update the score
            if not marks:
                marks = 0
            mark.score = marks
            mark.save()
            print(mark)

        messages.success(
            request,
            f"Marks have been updated for the subject: `{subject.name}` by: `{request.user.username}`",
        )
        return redirect(reverse("staff:admin-marks"))

    template_name = "students/upload-marks.html"
    context = {
        "section": "marks-area",
        "subject": subject,
        "klass": klass,
    }

    return render(request, template_name, context)
