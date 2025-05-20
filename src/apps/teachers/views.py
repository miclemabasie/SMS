import io
import json
from datetime import datetime, time, timezone
from random import random
import logging
from django.contrib import messages
from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required
from django.http import Http404, HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from faker import Faker
from openpyxl import load_workbook, workbook
import openpyxl
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.pdfgen import canvas
from reportlab.platypus import (
    Image,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from apps.announcements.models import Announcement, Event
from apps.fees.models import Fee
from apps.leave.models import TeacherLeave
from apps.students.forms import VerifyPinForm
from apps.students.models import (
    TEACHERSERVICE,
    Class,
    Department,
    Mark,
    StudentProfile,
    Subject,
    TeacherProfile,
    TeacherTempCreateProfile,
)
from apps.students.utils import (
    create_teacher_pin,
    generate_random_pin,
    get_student_temp_account,
    get_teacher_temp_account,
    send_account_creation_email,
    send_password_reset_email,
)
from apps.terms.models import AcademicYear, ExaminationSession, Term
from openpyxl.utils import get_column_letter

faker_factory = Faker()

User = get_user_model()
logger = logging.getLogger(__name__)


@login_required
def teacher_list_view(request, *args, **kwargs):

    teachers = TeacherProfile.objects.all()

    template_name = "teachers/teachers-list.html"

    context = {"section": "teachers-area", "teachers": teachers}

    return render(request, template_name, context)


@login_required
def teacher_detail_view(request, matricule, pkid, *args, **kwargs):
    teacher = get_object_or_404(TeacherProfile, matricule=matricule, pkid=pkid)
    template_name = "teachers/teacher-detail.html"
    pin = TeacherTempCreateProfile.objects.filter(teacher=teacher)
    # get all the subject tought by given teacher.
    assigned_subjects = teacher.assigned_subjects.all()
    if pin.exists():
        pin = pin.first()
    else:
        pin = None
    context = {
        "section": "teachers-area",
        "teacher": teacher,
        "pin": pin,
        "assigned_subjects": assigned_subjects,
    }

    return render(request, template_name, context)


@login_required
def teacher_edit_view(request, matricule, pkid, *args, **kwargs):
    teacher = get_object_or_404(TeacherProfile, matricule=matricule, pkid=pkid)

    if request.method == "POST":
        # Xtract data from the form
        first_name = request.POST.get("first_name")
        last_name = request.POST.get("last_name")
        subject = request.POST.get("subject")
        gender = request.POST.get("gender")
        phone = request.POST.get("phone")
        dob = request.POST.get("dob")
        address = request.POST.get("address")
        location = request.POST.get("location")
        country = request.POST.get("country")
        remark = request.POST.get("remark")
        username = request.POST.get("username")
        email = request.POST.get("email")
        profile_photo = request.FILES.get("photo")
        # extras
        region_of_origin = request.POST.get("region_of_origin")
        division_of_origin = request.POST.get("division_of_origin")
        sub_division_of_origin = request.POST.get("sub_division_of_origin")
        date_recruitement_public_service = request.POST.get(
            "date_recruitement_public_service"
        )
        corps = request.POST.get("corps")
        career_grade = request.POST.get("career_grade")
        payroll_grade = request.POST.get("payroll_grade")
        career_category = request.POST.get("career_category")
        payroll_category_solde = request.POST.get("payroll_category_solde")
        career_index = request.POST.get("career_index")
        payroll_index = request.POST.get("payroll_index")
        career_echelon = request.POST.get("career_echelon")
        payroll_echelon = request.POST.get("payroll_echelon")
        service = request.POST.get("service")
        appointed_structure = request.POST.get("appointed_structure")
        town = request.POST.get("town")
        possition_rank = request.POST.get("possition_rank")
        longivity_of_post = request.POST.get("longivity_of_post")
        longivity_in_administration = request.POST.get("longivity_in_administration")
        appointment_decision_reference = request.POST.get(
            "appointment_decision_reference"
        )
        indemnity_situation = request.POST.get("indemnity_situation")

        # Update user instance for the teacher

        teacher.user.username = username
        teacher.user.first_name = first_name
        teacher.user.last_name = last_name
        teacher.user.email = email

        teacher.user.save()

        # Update teacher profile instance

        teacher.gender = gender
        teacher.phone_number = phone
        teacher.main_subject = subject
        teacher.address = address

        teacher.save()
        if location:
            teacher.location = location
        if profile_photo:
            teacher.profile_photo = profile_photo
        if remark:
            teacher.remark = remark
        if country:
            teacher.country = country
        teacher.region_of_origin = region_of_origin
        teacher.division_of_origin = division_of_origin
        teacher.sub_division_of_origin = sub_division_of_origin
        teacher.date_recruitement_public_service = date_recruitement_public_service
        teacher.corps = corps
        teacher.career_grade = career_grade
        teacher.payroll_grade = payroll_grade
        teacher.career_category = career_category
        teacher.payroll_category_solde = payroll_category_solde
        teacher.career_index = career_index
        teacher.payroll_index = payroll_index
        teacher.career_echelon = career_echelon
        teacher.payroll_echelon = payroll_echelon
        teacher.service = service
        teacher.appointed_structure = appointed_structure
        teacher.town = town
        teacher.possition_rank = possition_rank
        teacher.longivity_of_post = longivity_of_post
        teacher.longivity_in_administration = longivity_in_administration
        teacher.appointment_decision_reference = appointment_decision_reference
        teacher.indemnity_situation = indemnity_situation

        # construct a valid date out of the html date
        if dob:
            date_string_from_form = dob

            dob = datetime.strptime(date_string_from_form, "%Y-%m-%d").date()
            # Create a specific time
            specific_time = time(8, 51, 2)

            dob_with_time = datetime.combine(dob, specific_time, tzinfo=timezone.utc)
            teacher.user.dob = dob_with_time

        teacher.save()
        teacher.user.save()

        messages.success(request, "TEacher successfully updated")

        return redirect(
            reverse(
                "teachers:teachers-detail",
                kwargs={"pkid": teacher.pkid, "matricule": teacher.matricule},
            )
        )

    template_name = "teachers/teacher-edit.html"
    context = {"teacher": teacher, "section": "teachers-area"}

    return render(request, template_name, context)


@login_required
def teacher_delete_view(request, matricule, pkid, *args, **kwargs):

    teacher = get_object_or_404(TeacherProfile, matricule=matricule, pkid=pkid)

    print("Deleting the user.")

    teacher.delete()

    messages.success(request, "Teacher successfuly deleted.")

    # template_name = "teachers/teacher-edit.html"

    return redirect(reverse("teachers:teachers-list"))


@login_required
def teacher_add_view(request, *args, **kwargs):

    if request.method == "POST":
        # Xtract data from the form
        first_name = request.POST.get("first_name")
        last_name = request.POST.get("last_name")
        subject = request.POST.get("subject")
        gender = request.POST.get("gender")
        phone = request.POST.get("phone")
        dob = request.POST.get("dob")
        address = request.POST.get("address")
        location = request.POST.get("location")
        country = request.POST.get("country")
        remark = request.POST.get("remark")
        username = request.POST.get("username")
        email = request.POST.get("email")
        profile_photo = request.FILES.get("photo")
        pin = request.POST.get("pin")

        # extras
        region_of_origin = request.POST.get("region_of_origin")
        division_of_origin = request.POST.get("division_of_origin")
        sub_division_of_origin = request.POST.get("sub_division_of_origin")
        date_recruitement_public_service = request.POST.get(
            "date_recruitement_public_service"
        )
        corps = request.POST.get("corps")
        career_grade = request.POST.get("career_grade")
        payroll_grade = request.POST.get("payroll_grade")
        career_category = request.POST.get("career_category")
        payroll_category_solde = request.POST.get("payroll_category_solde")
        career_index = request.POST.get("career_index")
        payroll_index = request.POST.get("payroll_index")
        career_echelon = request.POST.get("career_echelon")
        payroll_echelon = request.POST.get("payroll_echelon")
        service = request.POST.get("service")
        appointed_structure = request.POST.get("appointed_structure")
        town = request.POST.get("town")
        possition_rank = request.POST.get("possition_rank")
        longivity_of_post = request.POST.get("longivity_of_post")
        longivity_in_administration = request.POST.get("longivity_in_administration")
        appointment_decision_reference = request.POST.get(
            "appointment_decision_reference"
        )
        indemnity_situation = request.POST.get("indemnity_situation")

        if len(pin) == 4:
            # create temp
            pass
        else:
            messages.error(
                "Password reset pin has incorrect length, should be exactly 4."
            )
            return redirect(reverse("teachers:teachers-add"))

        # Create user instance for the teacher
        # construct a valid date out of the html date
        date_string_from_form = dob

        dob = datetime.strptime(date_string_from_form, "%Y-%m-%d").date()
        # Create a specific time
        specific_time = time(8, 51, 2)

        dob_with_time = datetime.combine(dob, specific_time, tzinfo=timezone.utc)

        user = User.objects.create(
            username=username,
            first_name=first_name,
            last_name=last_name,
            email=email,
            is_teacher=True,
            dob=dob_with_time,
        )
        user.save()
        user.set_password(faker_factory.password())
        user.save()

        # Create teacher profile instance

        teacher, created = TeacherProfile.objects.get_or_create(
            user=user,
            gender=gender,
            phone_number=phone,
            main_subject=subject,
            address=address,
        )
        if created:
            teacher.region_of_origin = region_of_origin
            teacher.division_of_origin = division_of_origin
            teacher.sub_division_of_origin = sub_division_of_origin
            teacher.date_recruitement_public_service = date_recruitement_public_service
            teacher.corps = corps
            teacher.career_grade = career_grade
            teacher.payroll_grade = payroll_grade
            teacher.career_category = career_category
            teacher.payroll_category_solde = payroll_category_solde
            teacher.career_index = career_index
            teacher.payroll_index = payroll_index
            teacher.career_echelon = career_echelon
            teacher.payroll_echelon = payroll_echelon
            teacher.service = service
            teacher.appointed_structure = appointed_structure
            teacher.town = town
            teacher.possition_rank = possition_rank
            teacher.longivity_of_post = longivity_of_post
            teacher.longivity_in_administration = longivity_in_administration
            teacher.appointment_decision_reference = appointment_decision_reference
            teacher.indemnity_situation = indemnity_situation

            if location:
                teacher.location = location
            if profile_photo:
                teacher.profile_photo = profile_photo
            if remark:
                teacher.remark = remark
            if country:
                teacher.country = country

            teacher.save()
            teacher.user.is_teacher = True
            teacher.user.save()
            create_teacher_pin(teacher, pin)
            send_account_creation_email(request, teacher.user, "teacher")
            return redirect(
                reverse(
                    "teachers:teachers-detail",
                    kwargs={"pkid": teacher.pkid, "matricule": teacher.matricule},
                )
            )
        else:
            messages.warning(request, "Teacher already exists.")
            return redirect(reverse("teachers:teachers-list"))

    template_name = "teachers/teacher-add.html"
    context = {"section": "teachers-area"}

    return render(request, template_name, context)


@login_required
def class_add_view(request, *args, **kwargs):

    if request.method == "POST":

        # extract information from the class form
        grade_level = request.POST.get("grade_level")
        class_name = request.POST.get("class_name")
        class_master = request.POST.get("class_master")
        class_prefect = request.POST.get("class_prefect")
        class_code = request.POST.get("class_code")
        department_pkid = request.POST.get("department")
        promotion_avg = request.POST.get("promotion_average")
        if department_pkid:
            department_pkid = int(department_pkid)
        try:
            department = Department.objects.get(pkid=department_pkid)
        except Department.DoesNotExist:
            messages.error(request, "Department Does not exist.")
            return redirect(reverse("class-list"))

        print(class_prefect, class_master)

        # Create the class
        if grade_level and class_name:

            klass = Class.objects.create(
                grade_level=grade_level,
                class_name=class_name,
                class_master=class_master,
                class_prefect=class_prefect,
                class_code=class_code,
            )
            if promotion_avg:
                promotion_avg = int(promotion_avg)
                klass.pass_avg = promotion_avg
            if department:
                klass.department = department
            klass.save()
        save_and_add_flag = request.POST.get("save_and_add")

        if save_and_add_flag:
            print(save_and_add_flag)
            messages.success(request, "Class Added")
            return redirect(reverse("class-add"))
        messages.success(request, "Class Added")
        return redirect(reverse("class-list"))
    departments = Department.objects.all()
    template_name = "classes/class-add.html"
    context = {"section": "class-area", "departments": departments}

    return render(request, template_name, context)


@login_required
def download_sample_class_file(request, *args, **kwargs):
    """
    View to download a sample Excel file for class uploads.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Class Upload Sample File"

    # Define the headers for the columns
    headers = [
        "Class Name",
        "Grade Level",
        "Class Master",
        "Class Prefect",
        "Class Code",
        "Promotion Average",
    ]

    # Write headers to the first row
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header

    # Set column widths for better readability
    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)
        ws.column_dimensions[column_letter].width = 25

    # Create an HTTP response object with the correct content type for Excel files
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    # Set the content disposition to attachment to trigger file download
    response["Content-Disposition"] = (
        "attachment; filename=class_upload_sample_file.xlsx"
    )

    # Save the workbook to the response object
    wb.save(response)

    return response


@login_required
def upload_classes_from_file(request, *args, **kwargs):
    """
    View to handle the file upload for class data.
    """
    if request.method == "POST" and request.FILES.get("classes_file"):
        # Get the uploaded file from the request
        classes_file = request.FILES["classes_file"]

        try:
            # Load the uploaded workbook
            wb = load_workbook(filename=classes_file)
            # Get the first sheet from the workbook
            ws = wb.active
        except Exception as e:
            # If there's an error reading the file, show an error message and redirect
            messages.error(request, f"Error reading the file: {e}")
            return redirect(reverse("classes:upload-classes-from-file"))

        # Iterate through each row in the worksheet starting from the second row
        for row_num, row in enumerate(
            ws.iter_rows(min_row=2, values_only=True), start=2
        ):
            # Extract values from the row
            (
                class_name,
                grade_level,
                class_master,
                class_prefect,
                class_code,
                promotion_avg,
            ) = row[:6]

            # Check if essential fields are present
            if not class_name or not grade_level:
                # If any essential field is missing, show an error message and redirect
                messages.error(
                    request,
                    f"Invalid file, missing essential information on row {row_num}",
                )
                return redirect(reverse("class-list"))

            try:
                # Create or update the Class object based on grade_level and class_name
                klass, created = Class.objects.update_or_create(
                    grade_level=grade_level,
                    class_name=class_name,
                    defaults={
                        "class_master": class_master,
                        "class_prefect": class_prefect,
                        "class_code": class_code,
                        "pass_avg": int(promotion_avg) if promotion_avg else None,
                    },
                )

                if created:
                    logger.info(f"Class {klass.class_name} created")
                else:
                    logger.info(f"Class {klass.class_name} updated")

            except Exception as e:
                # Log any errors that occur and show an error message
                logger.error(f"Error processing row {row_num}: {e}")
                messages.error(request, f"Error processing row {row_num}.")
                return redirect(reverse("class-list"))

        # Show a success message after processing all rows and redirect
        messages.success(request, "Classes have been uploaded successfully.")
        return redirect(reverse("class-list"))

    # Render the form template for uploading classes
    context = {"section": "class-area"}
    return redirect(reverse("class-list"))


@login_required
def class_list_view(request, *args, **kwargs):

    classes = Class.objects.all()

    template_name = "classes/class-list.html"
    context = {
        "section": "class-area",
        "classes": classes,
    }

    return render(request, template_name, context)


@login_required
def class_edit_view(request, pkid, *args, **kwargs):
    klass = get_object_or_404(Class, pkid=pkid)
    if request.method == "POST":

        # extract information from the class form
        grade_level = request.POST.get("grade_level")
        class_name = request.POST.get("class_name")
        class_master = request.POST.get("class_master")
        class_prefect = request.POST.get("class_prefect")
        promotion_avg = request.POST.get("promotion_average")
        class_code = request.POST.get("class_code")
        department_pkid = request.POST.get("department")
        if department_pkid:
            department_pkid = int(department_pkid)
        try:
            department = Department.objects.get(pkid=department_pkid)
        except Department.DoesNotExist:
            messages.error(request, "Deparment Does not exist.")
            return redirect(reverse("class-list"))

        print(class_prefect, class_master)

        klass.grade_level = grade_level
        klass.class_name = class_name
        klass.class_master = class_master
        klass.class_prefect = class_prefect
        klass.pass_avg = promotion_avg
        klass.class_code = class_code
        klass.department = department

        klass.save()
        save_and_add_flag = request.POST.get("save_and_add")
        if save_and_add_flag:
            print(save_and_add_flag)
            return redirect(reverse("class-add"))
        messages.success(request, "Updated class.")
        return redirect(reverse("class-list"))
    departments = Department.objects.all()
    template_name = "classes/class-edit.html"
    context = {"section": "class-area", "class": klass, "departments": departments}

    return render(request, template_name, context)


@login_required
def class_delete_view(request, pkid, *args, **kwargs):
    klass = get_object_or_404(Class, pkid=pkid)

    klass.delete()
    messages.success(request, "Class has been successfully deleted")

    return redirect(reverse("class-list"))


def verify_teacher_pin(request):
    if request.method == "POST":
        pin = request.POST.get("pin")
        email = request.POST.get("email")
        # Check if teacher exists with this email
        teachers = TeacherProfile.objects.filter(user__email=email)
        if teachers.exists():
            teacher = teachers.first()
        else:
            messages.error(request, "No teacher found with given email.")
            return redirect(reverse("users:user-login"))

        # Check if teacher temp profile exist with this information
        teacher_temp = get_teacher_temp_account(pin, email, teacher)

        if teacher_temp:
            # go ahead and send them the email to perform a reset
            send_password_reset_email(request, teacher.user, teacher_temp)
            return redirect(reverse("users:password_reset_done"))
        else:
            messages.error(request, "No pending setup for the given account")
            return redirect(reverse("users:user-login"))
    else:
        template_name = "accounts/verifypin-teacher.html"
        form = VerifyPinForm()
        context = {"form": form}

        return render(request, template_name, context)


@login_required
def teacher_dashboard(request):
    # check if current user is user of type teacher
    template_name = "dashboards/teachers/teacher-dashboard.html"
    if request.user.is_teacher or request.user.is_staff:
        announcements = Announcement.objects.filter(visible_to_students=True)
        teacher = None
        if request.user.is_teacher:
            teacher = request.user.teacher_profile

            assigned_subjects = Subject.objects.filter(assigned_to=teacher)

        # get the number of classes teacher is involved with
        klasses = []
        for sub in assigned_subjects:
            if sub.klass not in klasses:
                klasses.append(sub.klass)

        # Get all the leave for this teacher
        leaves = TeacherLeave.objects.filter(teacher=teacher).count()
        total_leave = 0
        if leaves > 0:
            total_leave = leaves

        # get all the events made for teachers
        events = Event.objects.filter(visible_to_teachers=True)

        context = {
            "announcements": announcements,
            "teacher": teacher,
            "assigned_subjects": assigned_subjects,
            "assigned_subjects_count": assigned_subjects.count(),
            "total_classes": len(klasses),
            "total_leaves": total_leave,
            "events": events,
            "section": "teacher-dashboard",
            # "total_pass_courses": pass_courses_count,
            # "total_cources_writen": marks.count(),
            # "total_courses": total_courses,
            # "payment_history": payment_history,
            # "class": student.current_class,
            # "class_enrolment": class_enrolment.count(),
        }
        return render(request, template_name, context)
    else:
        messages.error(request, "You are not allowed to be here.")
        return render(
            request, template_name, {"error_message": "You are not allowed to be here."}
        )


def assign_class_to_teacher(request, teacher_pkid, teacher_mat):
    teachers = TeacherProfile.objects.filter(pkid=teacher_pkid, matricule=teacher_mat)
    if teachers.exists():
        teacher = teachers.first()
    else:
        messages.error(request, "Teacher with given matricule, not found..")
        return redirect(reverse("staff:admin-dashboard"))
    # check if ajax was sent
    if request.method == "POST":
        # print("This are the selected subjects", request.body)
        data = json.loads(request.body)

        # get current assigned subject for the given teacher
        assigned_subjects = Subject.objects.filter(assigned_to=teacher)

        print(data)
        # extract all the IDs of the selected subjects and add them to list.
        selected_subjects_ids = []
        for subject in data["selectedSubjects"]:
            print(subject["pkid"])
            selected_subjects_ids.append(subject.get("pkid"))

        if len(selected_subjects_ids) < 1:
            # remove all the subject currently assigned to the teacher
            for sub in assigned_subjects:
                sub.assigned_to = None
                sub.assigned = False
                sub.save()
                return JsonResponse({"message": "updated.."})
        # Flush out all the subjects that exist in the  and reset

        if len(assigned_subjects) > 0 and len(selected_subjects_ids) > 0:
            for subject in assigned_subjects:
                # remove the subject
                subject.assigned_to = None
                subject.assigned = False
                subject.save()

        for pkid in selected_subjects_ids:
            sub = Subject.objects.filter(pkid=pkid).first()
            if sub:
                sub.assigned_to = teacher
                sub.assigned = True
                sub.save()

        return JsonResponse({"message": "updated.."})

    # continue if teacher exists....
    # get all subjects that are assigned to the current teacher
    assigned_subjects = Subject.objects.filter(assigned_to=teacher)
    # get remaining subjects...
    subjects = Subject.objects.filter(assigned=False)

    # compile list of unique subjects
    context = {
        "assigned_subjects": assigned_subjects,
        "subjects": subjects,
        "teacher": teacher,
    }
    template_name = "teachers/assign-subjects.html"
    return render(request, template_name, context)


@login_required
def download_blank_teacher_form(request):
    # Create a file-like buffer to receive PDF data
    buffer = io.BytesIO()

    # Create the PDF object, using the buffer as its "file."
    p = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4

    # Draw title
    p.setFont("Helvetica-Bold", 16)
    p.drawString(200, height - 50, "Teacher Information Form")

    # Draw teacher fields
    p.setFont("Helvetica", 12)
    y = height - 100

    fields = [
        "First Name",
        "Last Name",
        "Gender",
        "Phone Number",
        "Date of Birth",
        "Address",
        "Location",
        "Country",
        "Subject",
        "Region of Origin",
        "Division of Origin",
        "Sub Division of Origin",
        "Date of Recruitment into Public Service",
        "Corps",
        "Career Grade",
        "Payroll Grade",
        "Career Category",
        "Payroll Category Solde",
        "Career Index",
        "Payroll Index",
        "Career Echelon",
        "Payroll Echelon",
        "Service",
        "Appointed Structure",
        "Town",
        "Position Rank",
        "Longevity of Post",
        "Longevity in Administration",
        "Appointment Decision Reference",
        "Indemnity Situation",
        "Remark",
    ]

    max_label_width = max(
        [p.stringWidth(field + ":", "Helvetica", 12) for field in fields]
    )
    line_start_x = 50 + max_label_width + 10
    line_length = 300

    for field in fields:
        p.drawString(50, y, f"{field}:")
        p.line(line_start_x, y, line_start_x + line_length, y)
        y -= 20
        if y < 50:  # Check if y is less than the bottom margin
            p.showPage()
            p.setFont("Helvetica", 12)
            y = height - 50

    # Close the PDF object cleanly, and we're done.
    p.showPage()
    p.save()

    # Get the value of the BytesIO buffer and write it to the response.
    pdf = buffer.getvalue()
    buffer.close()

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="teacher_form.pdf"'
    response.write(pdf)

    return response


@login_required
def download_teacher_list(request, *args, **kwargs):
    if request.method == "GET":

        teachers = TeacherProfile.objects.all()

        wb = openpyxl.Workbook()
        ws = wb.active

        ws.title = "Teachers List"
        filename = "teachers-list.xlsx"

        headers = [
            "Matricule",
            "Fullname",
            "Gender",
            "Phone Number",
            "Country",
            "Location",
            "Address",
            "Main Subject",
            "Number of Absences",
            "Region of Origin",
            "Division of Origin",
            "Sub Division of Origin",
            "Date of Recruitment into Public Service",
            "Corps",
            "Career Grade",
            "Payroll Grade",
            "Career Category",
            "Payroll Category Solde",
            "Career Index",
            "Payroll Index",
            "Career Echelon",
            "Payroll Echelon",
            "Service",
            "Appointed Structure",
            "Town",
            "Position Rank",
            "Longevity of Post",
            "Longevity in Administration",
            "Reference of the Appointment Decision",
            "Indemnity Situation",
            "Remark",
        ]

        # Write headers to the first row
        for col_num, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header

        # Write teacher data to the worksheet
        for row_num, teacher in enumerate(teachers, start=2):
            ws.cell(row=row_num, column=1).value = teacher.matricule
            ws.cell(row=row_num, column=2).value = teacher.user.get_fullname
            ws.cell(row=row_num, column=3).value = teacher.gender
            ws.cell(row=row_num, column=4).value = str(teacher.phone_number)
            ws.cell(row=row_num, column=5).value = teacher.country.name
            ws.cell(row=row_num, column=6).value = teacher.location
            ws.cell(row=row_num, column=7).value = teacher.address
            ws.cell(row=row_num, column=8).value = teacher.main_subject
            ws.cell(row=row_num, column=9).value = teacher.number_of_absences
            ws.cell(row=row_num, column=10).value = teacher.region_of_origin
            ws.cell(row=row_num, column=11).value = teacher.division_of_origin
            ws.cell(row=row_num, column=12).value = teacher.sub_division_of_origin
            ws.cell(row=row_num, column=13).value = (
                teacher.date_recruitement_public_service
            )
            ws.cell(row=row_num, column=14).value = teacher.corps
            ws.cell(row=row_num, column=15).value = teacher.career_grade
            ws.cell(row=row_num, column=16).value = teacher.payroll_grade
            ws.cell(row=row_num, column=17).value = teacher.career_category
            ws.cell(row=row_num, column=18).value = teacher.payroll_category_solde
            ws.cell(row=row_num, column=19).value = teacher.career_index
            ws.cell(row=row_num, column=20).value = teacher.payroll_index
            ws.cell(row=row_num, column=21).value = teacher.career_echelon
            ws.cell(row=row_num, column=22).value = teacher.payroll_echelon
            ws.cell(row=row_num, column=23).value = teacher.service
            ws.cell(row=row_num, column=24).value = teacher.appointed_structure
            ws.cell(row=row_num, column=25).value = teacher.town
            ws.cell(row=row_num, column=26).value = teacher.possition_rank
            ws.cell(row=row_num, column=27).value = teacher.longivity_of_post
            ws.cell(row=row_num, column=28).value = teacher.longivity_in_administration
            ws.cell(row=row_num, column=29).value = (
                teacher.appointment_decision_reference
            )
            ws.cell(row=row_num, column=30).value = teacher.indemnity_situation
            ws.cell(row=row_num, column=31).value = teacher.remark

        # Set column widths
        for col_num in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_num)
            ws.column_dimensions[column_letter].width = 20

        # Create a response object
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f"attachment; filename={filename}"

        # Save the workbook to the response
        wb.save(response)

        return response

    return redirect(reverse("teachers:teacher-list"))


@login_required
def upload_teachers_from_file(request, *args, **kwargs):
    # Get all the available subjects
    subjects = Subject.objects.all()

    if request.method == "POST" and request.FILES.get("teachers_file"):
        teachers_file = request.FILES["teachers_file"]

        wb = load_workbook(filename=teachers_file)
        ws = wb.get_sheet_by_name("teachers")

        # Check if all the required attributes are available in the file
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            first_name, last_name, email, gender, phone, address = (
                row[0],
                row[1],
                row[2],
                row[3],
                row[4],
                row[5],
            )
            attributes = [first_name, last_name, gender, phone]
            for value in attributes:
                if not value:
                    messages.error(request, f"Invalid file, missing information on row {row_num}")
                    return redirect(reverse("teachers:upload-teachers-from-file"))

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            (
                first_name,
                last_name,
                email,
                gender,
                phone,
                address,
                country,
                location,
                main_subject,
                number_of_absences,
                region_of_origin,
                division_of_origin,
                sub_division_of_origin,
                date_recruitement_public_service,
                corps,
                career_grade,
                payroll_grade,
                career_category,
                payroll_category_solde,
                career_index,
                payroll_index,
                career_echelon,
                payroll_echelon,
                service,
                appointed_structure,
                town,
                possition_rank,
                longivity_of_post,
                longivity_in_administration,
                appointment_decision_reference,
                indemnity_situation,
                remark,
            ) = row

            print(
                f"firstname {first_name}, lastname {last_name}, email {email} gender {gender} phone {phone} address {address}"
            )

            # Handle cases where email might be missing or duplicated
            if not email:
                email = f"{first_name.lower()}{random.randint(1, 1000)}@gmail.com"
                print(f"Generated email: {email}")
            else:
                # Check if the email already exists
                if User.objects.filter(email=email).exists():
                    messages.error(request, f"User with email {email} already exists at row {row_num}.")
                    continue

            try:
                user, created = User.objects.get_or_create(
                    username=f"{first_name.lower()}{last_name.lower()}",
                    first_name=first_name,
                    last_name=last_name,
                    email=email,
                    is_teacher=True,
                )

                if created:
                    user.save()
                    print(f"User {user.username} was saved.")

                # Create or update the teacher profile
                gender = "Male" if gender.lower() in ["m", "male"] else "Female"
                teacher, created = TeacherProfile.objects.update_or_create(
                    user=user,
                    defaults={
                        "gender": gender,
                        "phone_number": str(phone),
                        "address": address or "",
                        "country": country or "CM",
                        "location": location or "",
                        "main_subject": main_subject or "",
                        "number_of_absences": number_of_absences or 0,
                        "region_of_origin": region_of_origin or "",
                        "division_of_origin": division_of_origin or "",
                        "sub_division_of_origin": sub_division_of_origin or "",
                        "date_recruitement_public_service": date_recruitement_public_service or "",
                        "corps": corps or "",
                        "career_grade": career_grade or "",
                        "payroll_grade": payroll_grade or "",
                        "career_category": career_category or "",
                        "payroll_category_solde": payroll_category_solde or "",
                        "career_index": career_index or "",
                        "payroll_index": payroll_index or "",
                        "career_echelon": career_echelon or "",
                        "payroll_echelon": payroll_echelon or "",
                        "service": service or TEACHERSERVICE.OTHER,
                        "appointed_structure": appointed_structure or "",
                        "town": town or "",
                        "possition_rank": possition_rank or "",
                        "longivity_of_post": longivity_of_post or "",
                        "longivity_in_administration": longivity_in_administration or "",
                        "appointment_decision_reference": appointment_decision_reference or "",
                        "indemnity_situation": indemnity_situation or "",
                        "remark": remark or "",
                    },
                )

                if created:
                    teacher.save()
                    send_account_creation_email(request, teacher.user)

            except Exception as e:
                logger.error(f"Error processing row {row_num}: {e}")
                messages.error(request, f"Error processing row {row_num}.")
                continue

        messages.success(request, "Teachers have been uploaded successfully.")
        return redirect(reverse("teachers:teachers-list"))

    template_name = "teachers/upload-teachers.html"
    context = {
        "section": "teachers-area",
        "subjects": subjects,
    }

    return render(request, template_name, context)


@login_required
def download_teacher_sample_file(request, *args, **kwargs):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Teacher Upload Sample File"
    filename = "teacher-upload-sample-file.xlsx"

    headers = [
        "First Name",
        "Last Name",
        "Email",
        "Gender",
        "Phone",
        "Address",
        "Country",
        "Location",
        "Main Subject",
        "Number of Absences",
        "Region of Origin",
        "Division of Origin",
        "Sub Division of Origin",
        "Date of Recruitment into Public Service",
        "Corps",
        "Career Grade",
        "Payroll Grade",
        "Career Category",
        "Payroll Category Solde",
        "Career Index",
        "Payroll Index",
        "Career Echelon",
        "Payroll Echelon",
        "Service",
        "Appointed Structure",
        "Town",
        "Position Rank",
        "Longevity of Post",
        "Longevity in Administration",
        "Reference of the Appointment Decision",
        "Indemnity Situation",
        "Remark",
    ]

    # Write headers to the first row
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header

    # Set column widths
    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)
        ws.column_dimensions[column_letter].width = 25

    # Create a response object
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f"attachment; filename={filename}"

    # Save the workbook to the response
    wb.save(response)

    return response


def generate_report_card(request):
    return
